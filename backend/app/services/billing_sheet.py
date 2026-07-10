"""Generate the Excel billing sheet: a Summary tab (schedule of values, linked to task
rows) plus one tab per invoice — exactly the artifact Joe assembles by hand today."""

import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models import Project
from app.services.review_engine import build_task_ledger, latest_contract

HEADER_FILL = PatternFill("solid", fgColor="00243A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="E8E8E8")
TOTAL_FONT = Font(bold=True)
WARNING_FILL = PatternFill("solid", fgColor="D6E408")
CRITICAL_FILL = PatternFill("solid", fgColor="F4A6A6")
CURRENCY_FMT = "$#,##0.00"
PCT_FMT = "0.0%"

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sheet_name(raw: str, used: set[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub("-", raw).strip()[:31] or "Sheet"
    base, i = name, 1
    while name in used:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _style_header(ws: Worksheet, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build_summary_sheet(wb: Workbook, project: Project) -> None:
    ws = wb.active
    ws.title = "Summary"

    contract = latest_contract(project)
    ws["A1"] = project.name
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Consultant: {project.consultant_name or '—'}"
    if contract:
        ws["A3"] = f"Contract: {contract.file_name} ({contract.label or 'n/a'})"
    else:
        ws["A3"] = "Contract: — (summary built from invoices)"

    header_row = 5
    headers = [
        "No.",
        "Cost Code",
        "Task Description",
        "Fee Type",
        "Contract Value",
        "Prior Billed",
        "Billed This Period",
        "Billed To Date",
        "% Billed",
        "Remaining",
        "Status",
    ]
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    _style_header(ws, header_row, len(headers))

    from app.services.review_engine import summarize_billing

    summary = summarize_billing(project)
    row = header_row + 1
    for r in summary.rows:
        ws.cell(row=row, column=1, value=r.task_number)
        ws.cell(row=row, column=2, value=r.cost_code)
        ws.cell(row=row, column=3, value=r.description)
        ws.cell(row=row, column=4, value="Lump Sum" if r.fee_type == "lump_sum" else "T&M")
        ws.cell(row=row, column=5, value=r.estimated_fee).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=r.prior_billed).number_format = CURRENCY_FMT
        ws.cell(row=row, column=7, value=r.billed_this_period).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8, value=r.billed_to_date).number_format = CURRENCY_FMT
        pct_cell = ws.cell(row=row, column=9, value=r.pct_billed)
        pct_cell.number_format = PCT_FMT
        ws.cell(row=row, column=10, value=r.remaining).number_format = CURRENCY_FMT
        status = "Not active" if not r.is_active else ("Over 100%" if r.flag_level == "critical" else ("75%+" if r.flag_level == "warning" else "OK"))
        status_cell = ws.cell(row=row, column=11, value=status)
        if r.flag_level == "critical":
            pct_cell.fill = CRITICAL_FILL
            status_cell.fill = CRITICAL_FILL
        elif r.flag_level == "warning":
            pct_cell.fill = WARNING_FILL
            status_cell.fill = WARNING_FILL
        if not r.is_active:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = Font(italic=True, color="999999")
        row += 1

    total_row = row
    ws.cell(row=total_row, column=3, value="CONTRACT TOTAL")
    ws.cell(row=total_row, column=5, value=summary.contract_total).number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=8, value=summary.total_billed_to_date).number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=10, value=summary.total_remaining).number_format = CURRENCY_FMT
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [6, 16, 46, 10, 14, 14, 16, 14, 10, 14, 12])


def _build_invoice_sheet(wb: Workbook, project: Project, invoice, used_names: set[str]) -> None:
    tasks_by_id = {t.id: t for t in (latest_contract(project).tasks if latest_contract(project) else [])}
    title = invoice.invoice_number or invoice.file_name.rsplit(".", 1)[0]
    ws = wb.create_sheet(_sheet_name(title, used_names))

    ws["A1"] = f"Invoice {invoice.invoice_number or ''} — {invoice.file_name}"
    ws["A1"].font = Font(bold=True, size=12)
    period = ""
    if invoice.period_start or invoice.period_end:
        period = f"{invoice.period_start or '?'} to {invoice.period_end or '?'}"
    ws["A2"] = f"Period: {period}    Invoice date: {invoice.invoice_date or '—'}    Total: ${invoice.total_amount or 0:,.2f}"

    header_row = 4
    headers = [
        "Task #",
        "Cost Code",
        "Description",
        "Date",
        "Person",
        "Qty",
        "Unit Rate",
        "Amount",
        "Category",
        "Matched Contract Task",
        "Flags",
    ]
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    _style_header(ws, header_row, len(headers))

    has_contract = bool(tasks_by_id)
    # Attribute each flag to its own row: line-level flags by line_item_id, contract-task flags by
    # task id. Flags with neither (invoice-wide, e.g. total mismatch) are listed below the table
    # rather than smeared onto every row.
    flags_by_line: dict[int, list[str]] = {}
    flags_by_task: dict[int, list[str]] = {}
    invoice_level_flags: list[str] = []
    for f in invoice.flags:
        label = f"[{f.severity.upper()}] {f.rule_code}"
        if f.line_item_id is not None:
            flags_by_line.setdefault(f.line_item_id, []).append(label)
        elif f.contract_task_id is not None:
            flags_by_task.setdefault(f.contract_task_id, []).append(label)
        else:
            invoice_level_flags.append(f"[{f.severity.upper()}] {f.rule_code}: {f.message}")

    row = header_row + 1
    for item in invoice.line_items:
        matched = tasks_by_id.get(item.contract_task_id)
        amount = item.billed_this_period if item.billed_this_period is not None else item.amount
        ws.cell(row=row, column=1, value=item.raw_task_number or (matched.task_number if matched else ""))
        ws.cell(row=row, column=2, value=item.raw_cost_code or (matched.cost_code if matched else ""))
        ws.cell(row=row, column=3, value=item.description)
        ws.cell(row=row, column=4, value=item.work_date.isoformat() if item.work_date else "")
        ws.cell(row=row, column=5, value=item.person_name or "")
        ws.cell(row=row, column=6, value=item.quantity)
        if item.unit_rate is not None:
            ws.cell(row=row, column=7, value=item.unit_rate).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8, value=amount).number_format = CURRENCY_FMT
        ws.cell(row=row, column=9, value=item.category)
        if matched:
            ws.cell(row=row, column=10, value=f"{matched.task_number} — {matched.description[:40]}")
        elif has_contract:
            # A contract exists but this line matched nothing — genuinely worth flagging.
            ws.cell(row=row, column=10, value="— unmatched —").fill = CRITICAL_FILL
        else:
            ws.cell(row=row, column=10, value="— (no contract) —")

        item_flags = list(flags_by_line.get(item.id, []))
        if item.contract_task_id is not None:
            item_flags += flags_by_task.get(item.contract_task_id, [])
        flag_cell = ws.cell(row=row, column=11, value="; ".join(item_flags))
        if any("CRITICAL" in f for f in item_flags):
            flag_cell.fill = CRITICAL_FILL
        elif item_flags:
            flag_cell.fill = WARNING_FILL
        row += 1

    if invoice_level_flags:
        row += 1
        ws.cell(row=row, column=1, value="Invoice-level flags:").font = Font(bold=True)
        for msg in invoice_level_flags:
            row += 1
            ws.cell(row=row, column=1, value=msg)

    ws.freeze_panes = f"A{header_row + 1}"
    _autosize(ws, [8, 14, 42, 12, 16, 8, 12, 14, 12, 34, 30])


def generate_billing_workbook(project: Project) -> Workbook:
    wb = Workbook()
    _build_summary_sheet(wb, project)
    used_names = {"Summary"}
    for invoice in sorted(project.invoices, key=lambda i: (i.period_start or i.uploaded_at.date(), i.id)):
        _build_invoice_sheet(wb, project, invoice, used_names)
    return wb
